import requests
from bs4 import BeautifulSoup

from openpyxl import load_workbook

work_book = load_workbook('ranking.xlsx')
work_sheet = work_book['geniemusic']

headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
data = requests.get('https://www.genie.co.kr/chart/top200?ditc=D&rtm=N&ymd=20190908',headers=headers)

soup = BeautifulSoup(data.text, 'html.parser')

#movie = soup.select('#old_content > table > tbody > tr')
#old_content > table > tbody > tr:nth-child(2) > td.title > div > a

musicrank = soup.select('#body-content > div.newest-list > div > table > tbody > tr')
#body-content > div.newest-list > div > table > tbody > tr:nth-child(1) > td.info > a.title.ellipsis


# movies (tr들) 의 반복문을 돌리기
rank = 1
row = 2
for music in musicrank:
    # movie 안에 a 가 있으면,
    # a_tag = movie.select_one('td.title > div > a')
    musictitle = music.select_one('td.info > a.title.ellipsis')
    artistname = music.select_one('a.artist.ellipsis')

    if not musictitle == None:
        # a의 text를 찍어본다.
        title = musictitle.text
        artist = artistname.text
        work_sheet.cell(row=row, column=1, value=rank)
        work_sheet.cell(row=row, column=2, value=title)
        work_sheet.cell(row=row, column=3, value=artist)
        rank += 1
        row += 1

work_book.save('ranking.xlsx')