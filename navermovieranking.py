import requests
from bs4 import BeautifulSoup

from openpyxl import load_workbook

work_book = load_workbook('ranking.xlsx')
work_sheet = work_book['navermovie']

headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
data = requests.get('https://movie.naver.com/movie/sdb/rank/rmovie.nhn?sel=pnt&date=20190909',headers=headers)

soup = BeautifulSoup(data.text, 'html.parser')

movies = soup.select('#old_content > table > tbody > tr')

# movies (tr들) 의 반복문을 돌리기
rank = 1
row = 2
for movie in movies:
    # movie 안에 a 가 있으면,
    a_tag = movie.select_one('td.title > div > a')
    if not a_tag == None:
        # a의 text를 찍어본다.
        title = a_tag.text
        star = movie.select('td.point')[0].text
        work_sheet.cell(row=row, column=1, value=rank)
        work_sheet.cell(row=row, column=2, value=title)
        work_sheet.cell(row=row, column=3, value=star)
        rank += 1
        row += 1

work_book.save('ranking.xlsx')